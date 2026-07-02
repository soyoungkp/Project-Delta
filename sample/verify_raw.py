"""
원본 엑셀 셀 기준 독립 검증 스크립트.

bq_to_long / compare_bq 로직을 쓰지 않고, openpyxl 로 두 파일을 직접 읽어
파일별로 (1) PABS CODE 행, (2) PABS 열 범위, (3) FWBS 데이터 구간을
자동 탐색한 뒤 FWBS + PABS + 실제 Cell Address 기준으로 물량을 비교한다.
"""

import re
import sys
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

# Windows 콘솔(cp949)에서도 한글/특수문자 깨지지 않도록 UTF-8 강제
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")

HERE = Path(__file__).parent
FWBS_PAT = re.compile(r"^C\d")          # 정상 FWBS: C + 숫자
PABS_PAT = re.compile(r"^A\d{1,2}$")    # PABS 코드: A01 ~ A20


def find_pabs_row(ws):
    """'PABS CODE' 라벨이 있는 행과 그 라벨 열을 찾는다."""
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str) and v.strip() == "PABS CODE":
                return r, c
    raise ValueError("'PABS CODE' 라벨을 찾을 수 없습니다.")


def find_pabs_cols(ws, pabs_row, label_col):
    """PABS CODE 행에서 A01~A20 값이 있는 열 -> 코드 매핑 (하드코딩 없음)."""
    mapping = {}  # col_idx -> 'A01'
    for c in range(label_col + 1, ws.max_column + 1):
        v = ws.cell(row=pabs_row, column=c).value
        if isinstance(v, str) and PABS_PAT.match(v.strip()):
            mapping[c] = v.strip()
    return mapping


def find_data_start(ws, pabs_row):
    """PABS 행 이후, 컬럼 B(FWBS)가 'C+숫자'로 시작하는 첫 행."""
    for r in range(pabs_row + 1, ws.max_row + 1):
        b = ws.cell(row=r, column=2).value
        if isinstance(b, str) and FWBS_PAT.match(b.strip()):
            return r
    raise ValueError("FWBS 데이터 시작행을 찾을 수 없습니다.")


def find_summary_start(ws, data_start):
    """하단 CCS/Area 요약표 시작행을 탐색 (FWBS 코드 중복 재등장 또는 'CCS' 마커)."""
    seen = set()
    for r in range(data_start, ws.max_row + 1):
        b = ws.cell(row=r, column=2).value
        if not isinstance(b, str):
            continue
        code = b.strip()
        if not code:
            continue
        if "CCS" in code or "Summary" in code:
            return r
        if code in seen:          # 이미 나온 FWBS가 다시 등장 -> 요약표 시작
            return r
        if FWBS_PAT.match(code):
            seen.add(code)
    return ws.max_row + 1         # 요약표 없음 -> 끝까지


def extract(path):
    """파일에서 (FWBS, PABS) -> {qty, cell} 추출 + 메타정보 반환."""
    ws = openpyxl.load_workbook(path, data_only=True).active
    pabs_row, label_col = find_pabs_row(ws)
    pabs_cols = find_pabs_cols(ws, pabs_row, label_col)
    data_start = find_data_start(ws, pabs_row)
    summary_start = find_summary_start(ws, data_start)

    data = {}  # (fwbs, pabs) -> (qty, cell_addr)
    for r in range(data_start, summary_start):
        b = ws.cell(row=r, column=2).value
        if not isinstance(b, str) or not FWBS_PAT.match(b.strip()):
            continue
        fwbs = b.strip()
        for c, pabs in pabs_cols.items():
            v = ws.cell(row=r, column=c).value
            qty = v if isinstance(v, (int, float)) else None
            cell = f"{get_column_letter(c)}{r}"
            data[(fwbs, pabs)] = (qty, cell)

    meta = {
        "sheet": ws.title,
        "pabs_row": pabs_row,
        "data_rows": (data_start, summary_start - 1),
        "pabs_cols": {get_column_letter(c): p for c, p in sorted(pabs_cols.items())},
        "col_range": (
            get_column_letter(min(pabs_cols)),
            get_column_letter(max(pabs_cols)),
        ),
        "pabs_list": sorted(set(pabs_cols.values())),
        "fwbs_count": len({k[0] for k in data}),
        "qty_sum": sum(q for q, _ in data.values() if q is not None),
    }
    return data, meta


def main():
    prev_name, rev_name = "Previous_BQ.xlsx", "Revised_BQ.xlsx"
    prev, pmeta = extract(HERE / prev_name)
    rev, rmeta = extract(HERE / rev_name)

    keys = set(prev) | set(rev)
    diffs = []
    for k in keys:
        pq, pcell = prev.get(k, (None, None))
        rq, rcell = rev.get(k, (None, None))
        a = pq if pq is not None else 0
        b = rq if rq is not None else 0
        if round(a - b, 6) != 0:
            diffs.append((k[0], k[1], pcell, rcell, pq, rq, round(b - a, 6)))

    print(f"비교 키 수(FWBS+PABS): {len(keys):,}")
    print(f"Difference 있는 항목: {len(diffs):,}\n")

    if diffs:
        diffs.sort(key=lambda x: abs(x[6]), reverse=True)
        print("=== Top 50 Differences (|diff| 내림차순) ===")
        print(f"{'FWBS':<12}{'PABS':<6}{'Prev@':<9}{'Rev@':<9}"
              f"{'Prev_Qty':>14}{'Rev_Qty':>14}{'Diff':>14}")
        for fwbs, pabs, pc, rc, pq, rq, d in diffs[:50]:
            print(f"{fwbs:<12}{pabs:<6}{str(pc):<9}{str(rc):<9}"
                  f"{_fmt(pq):>14}{_fmt(rq):>14}{d:>14,.4f}")
    else:
        print("=== 차이 0건 — 진단 정보 ===")
        print(f"Previous 파일: {prev_name}")
        print(f"Revised  파일: {rev_name}")
        for label, m in [("PREVIOUS", pmeta), ("REVISED", rmeta)]:
            print(f"\n[{label}]")
            print(f"  Sheet            : {m['sheet']}")
            print(f"  PABS CODE 행     : {m['pabs_row']}")
            print(f"  비교 행 범위     : {m['data_rows'][0]} ~ {m['data_rows'][1]}")
            print(f"  비교 열 범위     : {m['col_range'][0]} ~ {m['col_range'][1]}")
            print(f"  PABS 열 매핑     : {m['pabs_cols']}")
            print(f"  PABS 목록        : {m['pabs_list']}")
            print(f"  FWBS 개수        : {m['fwbs_count']:,}")
            print(f"  Quantity 합계    : {m['qty_sum']:,.4f}")
        print(f"\nPABS 목록 동일?   : {pmeta['pabs_list'] == rmeta['pabs_list']}")
        print(f"FWBS 개수 동일?   : {pmeta['fwbs_count'] == rmeta['fwbs_count']}")
        print(f"Qty 합계 동일?    : "
              f"{round(pmeta['qty_sum'] - rmeta['qty_sum'], 6) == 0}")

    # --- 기존 compare_bq 결과와 교차 확인 ---
    try:
        from compare_bq import compare_bq
        cmp = compare_bq(HERE / prev_name, HERE / rev_name)
        cmp_changes = int((cmp["Change_Type"] != "No Change").sum())
        print(f"\n[교차확인] compare_bq 변경건수={cmp_changes:,} | "
              f"raw 셀 차이건수={len(diffs):,} | "
              f"일치={cmp_changes == len(diffs)}")
    except Exception as e:  # noqa: BLE001
        print(f"\n[교차확인] compare_bq 비교 실패: {e!r}")


def _fmt(v):
    return "—" if v is None else f"{v:,.4f}"


if __name__ == "__main__":
    main()

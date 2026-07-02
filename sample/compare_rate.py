"""
Unit Rate / Cost Change 비교.

Quantity(L~AE)가 아니라 단가/원가 컬럼(AH~AO)을 FWBS 기준으로 비교한다.
컬럼은 하드코딩(AH~AO)하지 않고 헤더 행(PABS CODE 행 바로 위)의 라벨로 탐색한다.
Remark(AP) 같은 비-단가 컬럼은 자동 제외된다.

결과 컬럼:
    FWBS, Description, Unit, Compare_Item,
    Previous_Value, Revised_Value, Difference, Difference_Percent, Change_Type
"""

import re
from pathlib import Path

import numpy as np
import openpyxl
import pandas as pd

HERE = Path(__file__).parent
FWBS_PAT = re.compile(r"^C\d")
VAL_TOL = 6   # 값 비교/차이 반올림 자릿수

# 헤더(정규화) -> Compare_Item 매핑. 정규화 = 대문자 + 공백/개행 제거
HEADER_TO_ITEM = {
    "MHRSU/P": "Manhour U/P",
    "LABORU/P": "Labor U/P",
    "MATERIALU/P": "Material U/P",
    "EQU/P": "Equipment U/P",
    "T&CU/P": "Tool & Consume U/P",
    "INDIRECTU/P": "Indirect U/P",
    "TOTALU/P": "Total U/P",
    "TOTALCOST": "Total Cost",
}
# 출력 정렬용 표준 순서
ITEM_ORDER = list(HEADER_TO_ITEM.values())

DESC_COLS = range(3, 9)   # C~H (들여쓰기 계층 Description)
FWBS_COL = 2              # B
UNIT_COL = 11             # K


def _norm(s):
    return re.sub(r"\s+", "", str(s)).upper()


def _find_pabs_row(ws):
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str) and v.strip() == "PABS CODE":
                return r
    raise ValueError("'PABS CODE' 라벨을 찾을 수 없습니다.")


def _find_data_start(ws, pabs_row):
    for r in range(pabs_row + 1, ws.max_row + 1):
        b = ws.cell(row=r, column=FWBS_COL).value
        if isinstance(b, str) and FWBS_PAT.match(b.strip()):
            return r
    raise ValueError("FWBS 데이터 시작행을 찾을 수 없습니다.")


def _find_summary_start(ws, data_start):
    """하단 CCS/Area 요약표 시작행 (FWBS 재등장 또는 'CCS'/'Summary' 마커)."""
    seen = set()
    for r in range(data_start, ws.max_row + 1):
        b = ws.cell(row=r, column=FWBS_COL).value
        if not isinstance(b, str):
            continue
        code = b.strip()
        if not code:
            continue
        if "CCS" in code or "Summary" in code:
            return r
        if code in seen:
            return r
        if FWBS_PAT.match(code):
            seen.add(code)
    return ws.max_row + 1


def _find_cost_cols(ws, header_row):
    """헤더 행에서 단가/원가 컬럼만 {col_idx -> Compare_Item} 으로 반환."""
    mapping = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=header_row, column=c).value
        if v is None:
            continue
        item = HEADER_TO_ITEM.get(_norm(v))
        if item:
            mapping[c] = item
    missing = set(HEADER_TO_ITEM.values()) - set(mapping.values())
    if missing:
        raise ValueError(f"단가/원가 컬럼을 찾지 못했습니다: {sorted(missing)}")
    return mapping


def extract_rates(path):
    """파일에서 FWBS 단위 단가/원가를 Long 형태 records 로 추출."""
    ws = openpyxl.load_workbook(path, data_only=True).active
    pabs_row = _find_pabs_row(ws)
    header_row = pabs_row - 1
    data_start = _find_data_start(ws, pabs_row)
    summary_start = _find_summary_start(ws, data_start)
    cost_cols = _find_cost_cols(ws, header_row)

    records = []
    for r in range(data_start, summary_start):
        b = ws.cell(row=r, column=FWBS_COL).value
        if not isinstance(b, str) or not FWBS_PAT.match(b.strip()):
            continue
        fwbs = b.strip()

        desc = ""
        for di in DESC_COLS:                       # 가장 구체적인 레벨 텍스트
            dv = ws.cell(row=r, column=di).value
            if dv not in (None, ""):
                desc = str(dv).strip()
        uv = ws.cell(row=r, column=UNIT_COL).value
        unit = "" if uv is None else str(uv).strip()

        for c, item in cost_cols.items():
            v = ws.cell(row=r, column=c).value
            # blank/None/NaN/"" 및 비숫자는 모두 0 으로 정규화한다.
            # → 파일 내 모든 (FWBS, Item) 셀이 항상 숫자가 되므로, merge 후
            #   한쪽에만 없는 키는 'FWBS 자체가 그 파일에만 존재'하는 경우뿐이다.
            val = v if isinstance(v, (int, float)) and not pd.isna(v) else 0
            records.append(
                {"FWBS": fwbs, "Description": desc, "Unit": unit,
                 "Compare_Item": item, "Value": val}
            )
    return pd.DataFrame.from_records(records)


def compare_rate(prev_path, rev_path):
    prev = extract_rates(prev_path)
    rev = extract_rates(rev_path)

    keys = ["FWBS", "Compare_Item"]
    prev_v = prev[keys + ["Value"]].rename(columns={"Value": "Previous_Value"})
    rev_v = rev[keys + ["Value"]].rename(columns={"Value": "Revised_Value"})
    merged = prev_v.merge(rev_v, on=keys, how="outer")

    # 메타(Description/Unit)는 FWBS 단위 → Revised 우선, 없으면 Previous
    meta_cols = ["FWBS", "Description", "Unit"]
    meta = (
        rev[meta_cols].drop_duplicates("FWBS").set_index("FWBS")
        .combine_first(prev[meta_cols].drop_duplicates("FWBS").set_index("FWBS"))
        .reset_index()
    )
    merged = merged.merge(meta, on="FWBS", how="left")

    # 값은 추출 단계에서 0 으로 정규화됨. 따라서 NaN(=키 부재)은
    #   'FWBS 자체가 그 파일에만 존재'함을 의미한다 → Added/Deleted 판정 기준.
    p = merged["Previous_Value"]
    r = merged["Revised_Value"]
    has_p = p.notna()   # Previous 파일에 해당 FWBS 존재
    has_r = r.notna()   # Revised 파일에 해당 FWBS 존재

    # 둘 다 존재할 때만 값 비교(0 정규화 기준). blank↔0, 0↔0 등은 same=True → No Change
    same = np.isclose(p.fillna(0), r.fillna(0), rtol=0, atol=10 ** -VAL_TOL)
    conditions = [
        has_p & ~has_r,             # Previous 에만 있는 FWBS  → Deleted
        ~has_p & has_r,             # Revised 에만 있는 FWBS   → Added
        has_p & has_r & ~same,      # 둘 다 존재 + 값 다름      → Changed
    ]
    merged["Change_Type"] = np.select(
        conditions, ["Deleted", "Added", "Changed"], default="No Change"
    )

    merged["Difference"] = (r.fillna(0) - p.fillna(0)).round(VAL_TOL)
    denom = p.where(has_p & (p != 0))
    merged["Difference_Percent"] = merged["Difference"] / denom

    merged["Compare_Item"] = pd.Categorical(
        merged["Compare_Item"], categories=ITEM_ORDER, ordered=True
    )
    cols = [
        "FWBS", "Description", "Unit", "Compare_Item",
        "Previous_Value", "Revised_Value", "Difference",
        "Difference_Percent", "Change_Type",
    ]
    return (
        merged[cols]
        .sort_values(["FWBS", "Compare_Item"])
        .reset_index(drop=True)
    )


def main():
    result = compare_rate(HERE / "Previous_BQ.xlsx", HERE / "Revised_BQ.xlsx")

    full_csv = HERE / "BQ_rate_comparison_result.csv"
    result.to_csv(full_csv, index=False, encoding="utf-8-sig")

    counts = result["Change_Type"].value_counts()
    print("=== Unit Rate / Cost Change (Key = FWBS) ===")
    for t in ["Added", "Deleted", "Changed", "No Change"]:
        print(f"  {t:<10}: {int(counts.get(t, 0)):,}")
    print(f"  {'TOTAL':<10}: {len(result):,}")

    changes = result[result["Change_Type"] != "No Change"]
    changes_csv = HERE / "BQ_rate_changes_only.csv"
    changes.to_csv(changes_csv, index=False, encoding="utf-8-sig")

    print(f"\nSaved full result   -> {full_csv}")
    print(f"Saved changes only  -> {changes_csv}  ({len(changes):,} rows)")
    if not changes.empty:
        print("\nChanged 상위 예시 (Compare_Item 별 건수):")
        print(changes["Compare_Item"].value_counts().to_string())


if __name__ == "__main__":
    main()

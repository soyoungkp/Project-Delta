"""
Quantity / Rate-Cost 비교 결과를 하나의 Excel 리포트(bytes, V1.1)로 생성한다.

Sheet 구성 (이름 prefix 로 정렬):
    01_Summary
    02_Key_Changes        (Quantity + Rate/Cost 변경분 통합, Change_Category 포함)
    03_Quantity_Changes
    04_Rate_Cost_Changes  (Compare_Item, |Difference| 큰 순)
    05_Rate_Cost_by_Item  (Compare_Item 별 집계)
    06_Raw_Quantity_All
    07_Raw_Rate_Cost_All

모든 시트: Autofilter / Header Freeze / 컬럼 너비 / Change_Type 색상 / 숫자 서식.
비교 로직은 compare_bq / compare_rate 에 있으며 여기서는 가공/서식만 담당한다.
"""

import io
from datetime import datetime

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from compare_rate import ITEM_ORDER  # Compare_Item 표준 순서

# ---- 색상/서식 상수 ----------------------------------------------------------
FILLS = {
    "Added": PatternFill("solid", fgColor="C6EFCE"),    # green
    "Deleted": PatternFill("solid", fgColor="FFC7CE"),  # red
    "Changed": PatternFill("solid", fgColor="FFEB9C"),  # yellow
}
FONTS = {
    "Added": Font(color="006100"),
    "Deleted": Font(color="9C0006"),
    "Changed": Font(color="9C6500"),
}
HEADER_FILL = PatternFill("solid", fgColor="305496")
HEADER_FONT = Font(color="FFFFFF", bold=True)

NUM_FMT = "#,##0.00"
PCT_FMT = "0.00%"
INT_FMT = "#,##0"

MONEY_COLS = {"Previous_Qty", "Revised_Qty", "Previous_Value", "Revised_Value",
              "Difference", "Previous", "Revised", "Difference_Sum"}
PCT_COLS = {"Difference_Percent"}
INT_COLS = {"Added", "Deleted", "Changed", "No Change", "Total", "Count"}


def safe_int(value):
    """NaN / None / 빈 문자열은 0 으로, 그 외는 int 로 안전 변환."""
    if value is None:
        return 0
    if pd.isna(value):
        return 0
    if value == "":
        return 0
    return int(value)


# ---- 공용 계산 헬퍼 (app.py 와 공유) -----------------------------------------
def n_changes(df):
    """No Change 가 아닌(=실제 변경된) 행 수."""
    return int((df["Change_Type"] != "No Change").sum())


def revision_type(qty_changed, rate_changed):
    if qty_changed > 0 and rate_changed > 0:
        return "Quantity + Rate/Cost Revision"
    if qty_changed > 0:
        return "Quantity Revision"
    if rate_changed > 0:
        return "Rate/Cost Revision"
    return "No Revision Detected"


def changed_items(rate_df, top=None):
    """Compare_Item 별 변경 건수 (내림차순). top 지정 시 상위 N."""
    sub = rate_df[rate_df["Change_Type"] != "No Change"]
    vc = sub["Compare_Item"].value_counts()
    vc = vc[vc > 0]
    items = [(str(k), int(v)) for k, v in vc.items()]
    return items[:top] if top else items


def revision_comment(qty_df, rate_df):
    """Revision Type 기반 자동 코멘트."""
    q, r = n_changes(qty_df), n_changes(rate_df)
    rtype = revision_type(q, r)
    if rtype == "No Revision Detected":
        return "두 파일 간 Quantity / Rate·Cost 변경사항이 감지되지 않았습니다."
    tops = ", ".join(f"{name}({cnt})" for name, cnt in changed_items(rate_df, top=3))
    if rtype == "Rate/Cost Revision":
        return (f"물량(Quantity) 변경 없이 단가/원가만 변경되었습니다. "
                f"Rate/Cost 변경 {r}건. 주요 변경 항목: {tops}")
    if rtype == "Quantity Revision":
        return f"단가/원가 변경 없이 물량(Quantity)만 변경되었습니다. Quantity 변경 {q}건."
    return (f"물량과 단가/원가가 모두 변경되었습니다. "
            f"Quantity {q}건 / Rate·Cost {r}건. 주요 Rate/Cost 항목: {tops}")


# ---- 파생 테이블 -------------------------------------------------------------
def build_key_changes(qty_df, rate_df):
    """Quantity + Rate/Cost 변경분(No Change 제외) 통합 테이블."""
    q = qty_df[qty_df["Change_Type"] != "No Change"]
    r = rate_df[rate_df["Change_Type"] != "No Change"]
    q2 = pd.DataFrame({
        "Change_Category": "Quantity",
        "FWBS": q["FWBS"], "Description": q["Description"], "Unit": q["Unit"],
        "Item": q["PABS"], "Previous": q["Previous_Qty"], "Revised": q["Revised_Qty"],
        "Difference": q["Difference"], "Difference_Percent": q["Difference_Percent"],
        "Change_Type": q["Change_Type"],
    })
    r2 = pd.DataFrame({
        "Change_Category": "Rate/Cost",
        "FWBS": r["FWBS"], "Description": r["Description"], "Unit": r["Unit"],
        "Item": r["Compare_Item"].astype(str), "Previous": r["Previous_Value"],
        "Revised": r["Revised_Value"], "Difference": r["Difference"],
        "Difference_Percent": r["Difference_Percent"], "Change_Type": r["Change_Type"],
    })
    return pd.concat([q2, r2], ignore_index=True)


def build_rate_by_item(rate_df):
    """Compare_Item 별 Added/Deleted/Changed/No Change/Total + Difference 합계."""
    rows = []
    for item in ITEM_ORDER:
        sub = rate_df[rate_df["Compare_Item"] == item]
        vc = sub["Change_Type"].value_counts()
        rows.append({
            "Compare_Item": item,
            "Added": safe_int(vc.get("Added", 0)),
            "Deleted": safe_int(vc.get("Deleted", 0)),
            "Changed": safe_int(vc.get("Changed", 0)),
            "No Change": safe_int(vc.get("No Change", 0)),
            "Total": len(sub),
            "Difference_Sum": round(float(sub["Difference"].sum()), 4),
        })
    return pd.DataFrame(rows)


def sort_rate_changes(rate_df):
    """Rate/Cost 변경분을 Compare_Item, |Difference| 큰 순으로 정렬."""
    rc = rate_df[rate_df["Change_Type"] != "No Change"].copy()
    if rc.empty:
        return rc
    rc["_abs"] = rc["Difference"].abs()
    rc = rc.sort_values(["Compare_Item", "_abs"], ascending=[True, False])
    return rc.drop(columns="_abs")


# ---- 시트 서식 ---------------------------------------------------------------
def _style_sheet(ws, df):
    ncol = len(df.columns)
    for c in range(1, ncol + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # 컬럼 너비 (앞쪽 200행 샘플)
    for c, name in enumerate(df.columns, start=1):
        vals = df[name].head(200).tolist()
        maxlen = max([len(str(name))] + [len(str(v)) for v in vals])
        width = 50 if name == "Description" else min(max(maxlen + 2, 10), 40)
        ws.column_dimensions[get_column_letter(c)].width = width

    # 숫자 서식
    max_row = ws.max_row
    for c, name in enumerate(df.columns, start=1):
        if name in MONEY_COLS:
            fmt = NUM_FMT
        elif name in PCT_COLS:
            fmt = PCT_FMT
        elif name in INT_COLS:
            fmt = INT_FMT
        else:
            continue
        for r in range(2, max_row + 1):
            ws.cell(row=r, column=c).number_format = fmt

    # Change_Type 색상 (No Change 제외 → 대용량에서 빠름)
    if "Change_Type" in df.columns:
        for i, ct in enumerate(df["Change_Type"].tolist()):
            fill = FILLS.get(ct)
            if fill is None:
                continue
            font = FONTS[ct]
            xr = i + 2
            for c in range(1, ncol + 1):
                ws.cell(row=xr, column=c).fill = fill
                ws.cell(row=xr, column=c).font = font


def _counts(df):
    vc = df["Change_Type"].value_counts()
    out = {t: safe_int(vc.get(t, 0)) for t in ["Added", "Deleted", "Changed", "No Change"]}
    out["Total"] = len(df)
    return out


def _add_sheet(writer, name, df, empty_msg):
    """데이터 시트 추가. 비어 있으면 안내 문구만 표시."""
    wb = writer.book
    if df is None or df.empty:
        ws = wb.create_sheet(name)
        ws["A1"] = empty_msg
        ws["A1"].font = Font(bold=True, color="9C0006")
        ws.column_dimensions["A"].width = 40
        return
    df.to_excel(writer, sheet_name=name, index=False)
    _style_sheet(wb[name], df)


def _write_summary(ws, prev_name, rev_name, qty_df, rate_df, generated):
    bold = Font(bold=True)
    title = Font(bold=True, size=14)
    section = Font(bold=True, color="305496")

    ws["A1"] = "Project Delta - BQ Revision Comparison Report (V1.1)"
    ws["A1"].font = title

    q, r = n_changes(qty_df), n_changes(rate_df)
    rtype = revision_type(q, r)

    r_idx = 3
    for label, val in [("Generated Time", generated),
                       ("Previous File Name", prev_name),
                       ("Revised File Name", rev_name),
                       ("Revision Type", rtype),
                       ("Quantity Changed Count", q),
                       ("Rate/Cost Changed Count", r)]:
        ws.cell(row=r_idx, column=1, value=label).font = bold
        ws.cell(row=r_idx, column=2, value=val)
        r_idx += 1

    # 요약 카운트 표
    order = ["Added", "Deleted", "Changed", "No Change", "Total"]
    for sect_title, df in [("Quantity Summary", qty_df),
                           ("Rate/Cost Summary", rate_df)]:
        r_idx += 1
        ws.cell(row=r_idx, column=1, value=sect_title).font = section
        r_idx += 1
        ws.cell(row=r_idx, column=1, value="Metric").font = bold
        ws.cell(row=r_idx, column=2, value="Count").font = bold
        c = _counts(df)
        for k in order:
            r_idx += 1
            ws.cell(row=r_idx, column=1, value=k)
            cell = ws.cell(row=r_idx, column=2, value=c[k])
            cell.number_format = INT_FMT
            if k in FILLS:
                cell.fill = FILLS[k]
                cell.font = FONTS[k]

    # Major Changed Compare Items
    r_idx += 2
    ws.cell(row=r_idx, column=1, value="Major Changed Compare Items").font = section
    r_idx += 1
    # changed_items() 는 [Compare_Item, Changed_Count] DataFrame 또는 None 반환
    tops = changed_items(rate_df, top=5)
    if tops is not None and not tops.empty:
        cnt_col = "Changed_Count" if "Changed_Count" in tops.columns else tops.columns[-1]
        ws.cell(row=r_idx, column=1, value="Compare_Item").font = bold
        ws.cell(row=r_idx, column=2, value="Changed").font = bold
        for _, row in tops.iterrows():
            r_idx += 1
            ws.cell(row=r_idx, column=1, value=str(row["Compare_Item"]))
            ws.cell(row=r_idx, column=2,
                    value=safe_int(row[cnt_col])).number_format = INT_FMT
    else:
        ws.cell(row=r_idx, column=1, value="(없음)")

    # 자동 코멘트
    r_idx += 2
    ws.cell(row=r_idx, column=1, value="Comment").font = section
    r_idx += 1
    ws.cell(row=r_idx, column=1, value=revision_comment(qty_df, rate_df)).alignment = (
        Alignment(wrap_text=True, vertical="top")
    )

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 60


def build_excel_report(qty_df, rate_df, prev_name, rev_name, generated=None,
                       include_raw=False):
    """비교 결과 두 DataFrame -> V1.1 Excel 리포트 bytes.

    include_raw=False(기본): 01~05 시트만 포함 (대용량 Raw 시트 제외 → 빠름).
    include_raw=True       : 06_Raw_Quantity_All / 07_Raw_Rate_Cost_All 추가.
    """
    generated = generated or datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    qty_changes = qty_df[qty_df["Change_Type"] != "No Change"]
    key_changes = build_key_changes(qty_df, rate_df)
    rate_changes_sorted = sort_rate_changes(rate_df)
    rate_by_item = build_rate_by_item(rate_df)

    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        _add_sheet(writer, "02_Key_Changes", key_changes,
                   "No Changes Detected")
        _add_sheet(writer, "03_Quantity_Changes", qty_changes,
                   "No Quantity Changes Detected")
        _add_sheet(writer, "04_Rate_Cost_Changes", rate_changes_sorted,
                   "No Rate/Cost Changes Detected")
        _add_sheet(writer, "05_Rate_Cost_by_Item", rate_by_item, "No Data")
        if include_raw:   # 대용량 Raw 시트는 옵션 선택 시에만 (생성 속도 영향 큼)
            _add_sheet(writer, "06_Raw_Quantity_All", qty_df, "No Data")
            _add_sheet(writer, "07_Raw_Rate_Cost_All", rate_df, "No Data")

        wb = writer.book
        ws_summary = wb.create_sheet("01_Summary")
        _write_summary(ws_summary, prev_name, rev_name, qty_df, rate_df, generated)

        # 이름 prefix(01_, 02_ ...) 기준 정렬로 시트 순서 확정
        wb._sheets.sort(key=lambda s: s.title)

    buffer.seek(0)
    return buffer.getvalue()



def n_changes(df):
    """Count rows with Added / Deleted / Changed."""
    if df is None or df.empty:
        return 0

    if "Change_Type" not in df.columns:
        return 0

    return int(df["Change_Type"].isin(["Added", "Deleted", "Changed"]).sum())


def revision_type(qty_df, rate_df):
    """Return revision type based on Quantity and Rate/Cost changes."""
    qty_changed = n_changes(qty_df)
    rate_changed = n_changes(rate_df)

    if qty_changed > 0 and rate_changed > 0:
        return "Quantity + Rate/Cost Revision"
    elif qty_changed > 0 and rate_changed == 0:
        return "Quantity Revision"
    elif qty_changed == 0 and rate_changed > 0:
        return "Rate/Cost Revision"
    else:
        return "No Revision Detected"


def changed_items(rate_df):
    """Return changed count by Compare_Item."""
    if rate_df is None or rate_df.empty:
        return None

    if "Change_Type" not in rate_df.columns or "Compare_Item" not in rate_df.columns:
        return None

    changed = rate_df[rate_df["Change_Type"].isin(["Added", "Deleted", "Changed"])]

    if changed.empty:
        return None

    return (
        changed.groupby("Compare_Item")
        .size()
        .reset_index(name="Changed_Count")
        .sort_values("Changed_Count", ascending=False)
    )


def revision_comment(qty_df, rate_df):
    """Generate short revision comment."""
    qty_changed = n_changes(qty_df)
    rate_changed = n_changes(rate_df)

    if qty_changed == 0 and rate_changed == 0:
        return "No quantity or rate/cost changes detected."

    if qty_changed == 0 and rate_changed > 0:
        return "Quantity 변경은 없으며, Unit Rate / Cost 변경만 발생했습니다."

    if qty_changed > 0 and rate_changed == 0:
        return "Quantity 변경만 발생했으며, Unit Rate / Cost 변경은 없습니다."

    return "Quantity 변경과 Unit Rate / Cost 변경이 모두 발생했습니다."

# ---------------------------------------------------------------------
# Helper functions for Streamlit UI
# These functions accept either DataFrame or already-counted integer.
# ---------------------------------------------------------------------

def n_changes(x):
    """Count Added / Deleted / Changed rows from DataFrame or int."""
    if x is None:
        return 0

    if isinstance(x, (int, float)):
        return safe_int(x)

    if hasattr(x, "empty"):
        if x.empty:
            return 0

        if "Change_Type" not in x.columns:
            return 0

        return safe_int(x["Change_Type"].isin(["Added", "Deleted", "Changed"]).sum())

    return 0


def revision_type(qty_input, rate_input):
    """Return revision type based on Quantity and Rate/Cost changes."""
    qty_changed = n_changes(qty_input)
    rate_changed = n_changes(rate_input)

    if qty_changed > 0 and rate_changed > 0:
        return "Quantity + Rate/Cost Revision"
    elif qty_changed > 0 and rate_changed == 0:
        return "Quantity Revision"
    elif qty_changed == 0 and rate_changed > 0:
        return "Rate/Cost Revision"
    else:
        return "No Revision Detected"


def changed_items(rate_df):
    """Return changed count by Compare_Item."""
    if rate_df is None:
        return None

    if not hasattr(rate_df, "empty"):
        return None

    if rate_df.empty:
        return None

    if "Change_Type" not in rate_df.columns or "Compare_Item" not in rate_df.columns:
        return None

    changed = rate_df[rate_df["Change_Type"].isin(["Added", "Deleted", "Changed"])]

    if changed.empty:
        return None

    return (
        changed.groupby("Compare_Item")
        .size()
        .reset_index(name="Changed_Count")
        .sort_values("Changed_Count", ascending=False)
    )


def revision_comment(qty_input, rate_input):
    """Generate short revision comment."""
    qty_changed = n_changes(qty_input)
    rate_changed = n_changes(rate_input)

    if qty_changed == 0 and rate_changed == 0:
        return "No quantity or rate/cost changes detected."

    if qty_changed == 0 and rate_changed > 0:
        return "Quantity 변경은 없으며, Unit Rate / Cost 변경만 발생했습니다."

    if qty_changed > 0 and rate_changed == 0:
        return "Quantity 변경만 발생했으며, Unit Rate / Cost 변경은 없습니다."

    return "Quantity 변경과 Unit Rate / Cost 변경이 모두 발생했습니다."


def changed_items(rate_df, top=None):
    """Return changed count by Compare_Item. Supports optional top argument."""
    if rate_df is None:
        return None

    if not hasattr(rate_df, "empty"):
        return None

    if rate_df.empty:
        return None

    if "Change_Type" not in rate_df.columns or "Compare_Item" not in rate_df.columns:
        return None

    changed = rate_df[rate_df["Change_Type"].isin(["Added", "Deleted", "Changed"])]

    if changed.empty:
        return None

    result = (
        changed.groupby("Compare_Item")
        .size()
        .reset_index(name="Changed_Count")
        .sort_values("Changed_Count", ascending=False)
    )

    if top is not None:
        result = result.head(top)

    return result
